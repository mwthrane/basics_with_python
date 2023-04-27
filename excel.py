import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
prod_list = inv_file["Sheet1"]

prod_per_supplier = {}
print(prod_list.max_row)

for prod_row in range(2, prod_list.max_row +1)
    supplier_name = prod_list.cell